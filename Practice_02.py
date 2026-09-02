from openpyxl import load_workbook

workbook = load_workbook("shipments.xlsx")

sheet = workbook.active

critical = 0
total_pallets = 0

for row in sheet.iter_rows(min_row=2):
        number = row[0].value
        customer = row[1].value
        delay = row[2].value
        pallets = row[3].value
        total_pallets += pallets

        if delay > 2:
            critical += 1
            print(f"Shipment number: {number}")
            print(f"Customer: {customer}")
            print(f"Delay: {delay}")
            print(f"Status: critical delay")
            print(f"Critical shipments: {critical}")
            print(f"Pallets: {pallets}")

#     if delay == 0:
#        status = "No delay"
#     elif delay <= 2:
#        status = "Small delay" 
#     else:
#        status = "Critical delay"
#        critical += 1

#     print(f"Shipment number: {number}")
#     print(f"Customer: {customer}")
#     print(f"Delay: {delay}")
#     print(f"Status: {status}")
#     print(f"Critical shipments: {critical}")
#     print(f"Pallets: {pallets}")
#     print()
# print()    
# print(f"Total pallets: {total_pallets}")    

print()   
print(f"Total pallets: {total_pallets}")    

