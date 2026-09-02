from openpyxl import load_workbook
from openpyxl import Workbook

source = load_workbook("shipments.xlsx")
source_sheet = source.active

report = Workbook()
report_sheet = report.active

report_sheet["A1"] = "Shipment"
report_sheet["B1"] = "Customer"
report_sheet["C1"] = "Delay"
report_sheet["D1"] = "Pallets"
report_row = 2

for row in source_sheet.iter_rows(min_row=2):
    number = row[0].value
    customer = row[1].value
    delay = row[2].value
    pallets = row[3].value
    if delay > 2:
        report_sheet[f"A{report_row}"] = number
        report_sheet[f"B{report_row}"] = customer
        report_sheet[f"C{report_row}"] = delay
        report_sheet[f"D{report_row}"] = pallets
        report_row += 1

report.save("criticals.xlsx")         
    
