from openpyxl import load_workbook

workbook = load_workbook("shipments.xlsx")

sheet = workbook.active 

for row in sheet.iter_rows(min_row=2):
    number = row[0].value
    customer = row[1].value
    delay = row[2].value

    print(f"Shipment: {number}")
    print(f"Customer: {customer}")
    print(f"Delay: {delay}")
    print()
