from openpyxl import load_workbook
from openpyxl import Workbook

source = load_workbook("shipments.xlsx")
sheet_source = source.active

high_priority = Workbook()
sheet_high_priority = high_priority.active

sheet_high_priority["A1"] = "Shipment"
sheet_high_priority["B1"] = "Customer"
sheet_high_priority["C1"] = "Delay"
sheet_high_priority["D1"] = "Priority"
report_row = 2
def get_priority(delay):
    if 5 <= delay <= 6:
        return "High"
    
    elif delay >= 7:
        return "Very High"
    else:
        return False

for row in sheet_source.iter_rows(min_row=2):
    number = row[0].value
    customer = row[1].value
    delay = row[2].value
    pallets = row[3].value
    priority = get_priority(delay)

    if priority:

        sheet_high_priority[f"A{report_row}"] = number
        sheet_high_priority[f"B{report_row}"] = customer
        sheet_high_priority[f"C{report_row}"] = delay
        sheet_high_priority[f"D{report_row}"] = priority
        report_row += 1
    

high_priority.save("high_priority.xlsx")



    


