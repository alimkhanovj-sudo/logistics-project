from openpyxl import load_workbook
from openpyxl import Workbook
source = load_workbook("shipments.xlsx")
sheet_source = source.active
summary = Workbook()
sheet_summary = summary.active

sheet_summary["A1"] = "Metric"
sheet_summary["B1"] = "Value"

sheet_summary["A2"] = "Total shipments"
sheet_summary["A3"] = "Critical shipments"
sheet_summary["A4"] = "Total pallets"
sheet_summary["A5"] = "Average delay"

total_shipments = 0
critical_shipments = 0
total_pallets = 0
total_delay = 0
average_delay = 0

def critical(delay):
    if delay >= 5:
        return True
    else: 
        return False
for row in sheet_source.iter_rows(min_row=2):
    delay = row[2].value
    pallets = row[3].value
    total_shipments += 1
    total_pallets += pallets
    total_delay += delay
   
    if critical(delay):
        critical_shipments += 1
average_delay = total_delay/total_shipments 

sheet_summary["B2"] = total_shipments     
sheet_summary["B3"] = critical_shipments    
sheet_summary["B4"] = total_pallets    
sheet_summary["B5"] = average_delay  
summary.save("summary.xlsx")  
       
