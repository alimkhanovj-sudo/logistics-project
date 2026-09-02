from openpyxl import load_workbook

def get_status(delay):
    if delay <= 2:
        return "Small delay"
    elif delay == 0:
        return "No delay"
    else:
        return "Critical delay"

workbook = load_workbook("shipments.xlsx")

sheet = workbook.active

for row in sheet.iter_rows(min_row=2):
    number = row[0].value
    customer = row[1].value
    delay = row[2].value
    pallets = row[3].value

    status = get_status(delay)
    print(f"Shipment number: {number}")
    print(f"Customer: {customer}")
    print(f"Delay: {delay}")
    print(f"Status: {status}")
    print()



# def is_critical(delay):
#     if delay > 2:
#         return True
    
#     else:
#         return False
    
# workbook = load_workbook("shipments.xlsx")
# sheet = workbook.active

# for row in sheet.iter_rows(min_row=2):
#     number = row[0].value
#     customer = row[1].value
#     delay = row[2].value
#     pallets = row[3].value
#     if is_critical(delay):
#         print(f"Critical shipment: {number}")
#     else:
#         print(f"No Critical: {number}")    
